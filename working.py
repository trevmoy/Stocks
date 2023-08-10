"""
Testing the import openpyxl module
"""

# Imports from openpyxl and openpyxl.utils
from stocks import stocks
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import *
import pandas as pd
from openpyxl.chart import LineChart, Reference, Series
import matplotlib.pyplot as plt


def main():
    """
    In main the input and function calls occur
    """

    # Creates the workbook and worksheet to variables wb and ws
    wb = Workbook()
    ws = wb.active
    correct_info: bool = False
    see_nasdaq: str 
    stock_dataframe: pd

    # Asks user if they want to see the nasdaq listings
    print("Would you like to see the tickers listed on the nasdaq?\n(Y/N)")
    see_nasdaq = input("")
    if (see_nasdaq == "Y" or see_nasdaq == "y"):
        print(stocks.get_nasdaq())

    # Object instantiation is placed in a try-except handler to catch invalid input
    while(correct_info == False):
        """
        The while loop will continue to ask for input inside of a
        try-except-else condition to handle any errors caused by incorrect user input.
        The try block asks for the 4 things required from the user and then creates an object called user with them.
        It then tests to see if the function returns a 
        pandas df type and if there is a manual error is raised and the exception repeats
        the loop. When the input is valid, the loop and try-except-else will end.
        """
        try:
            ticker = input("Please enter a ticker for a company: ")
            startDate = input("Please enter a start date(ex: 12/06/2005): ")
            endDate = input("Please enter an end date(ex: 12/06/2005): ")
            interval = input("Please enter the interval you'd like(ex: 1d, 1wk, 1mo): ")
            user = stocks(ticker, startDate, endDate, interval)
            
            if isinstance(user.get_info(), pd.DataFrame):
                print("\nThis information works\n")
                stock_dataframe = user.get_info()
                correct_info = True
            else:
                raise AssertionError

            
        except:
            print("\nYou've entered invalid information, please re-enter\n")
            correct_info = False
        else:
            correct_info = True

    # Sorts through df retrieved from the stocks class, sets index & header equal to true and appends each cell with the df info
    for r in dataframe_to_rows(user.get_info(), index=True, header=True):
        ws.append(r)

    # Corrects the sheet by deleting the unecessary empty space
    ws.delete_cols(1, 1)
    ws.delete_rows(2, 2)
    
    # Sets the index and header style to the default pandas style
    for cell in ws[1]:
        cell.style = 'Pandas'

    # Adjusts column width for date column
    col_len = len(str(ws['A4'].value))
    ws.column_dimensions['A'].width = col_len

    # Adjusts column width for volume column
    ws.column_dimensions['G'].width = 12

    # Attempting to plot data on a graph
    plt.style.use('ggplot')

    print(type(stock_dataframe)) 
    stock_dataframe.plot()

    # values = Reference(ws, min_col=0, min_row=0, max_col=len(ws['1']), max_row=len(ws['open']))
    # chart = LineChart()
    # chart.add_data(values)
    # ws.add_chart(chart, 'J15')

    # Asks user for file name
    file_name = input("Please enter a file name for your excell sheet: ")
    # Saves the document at the end of the code to update the actual excel sheet
    wb.save(f'{file_name.strip()}.xlsx')
    
    # Print statement to let users know that the spreadsheet is ready for viewing
    print("\nYour new excel sheet is ready!")

main()
